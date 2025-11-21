import tkinter as tk
from tkinter import ttk
from openpyxl import Workbook, load_workbook
from openpyxl.chart import BarChart, Reference
from datetime import datetime
import os
import json

STATE_FILE = "timewarp_state.json"
entrata = None

# Genera il nome del file con timestamp del giorno
def get_file_name(data=None):
    if data is None:
        data = datetime.now()
    elif isinstance(data, str):
        data = datetime.strptime(data, "%Y-%m-%d")
    return f"orari_lavoro_{data.strftime('%Y%m%d')}.xlsx"

# Carica stato entrata da file se esiste
def carica_stato():
    global entrata
    if os.path.exists(STATE_FILE):
        try:
            with open(STATE_FILE, 'r') as f:
                data = json.load(f)
                if data.get('entrata'):
                    entrata = datetime.fromisoformat(data['entrata'])
        except Exception:
            pass

# Salva stato entrata su file
def salva_stato():
    try:
        with open(STATE_FILE, 'w') as f:
            data = {'entrata': entrata.isoformat() if entrata else None}
            json.dump(data, f)
    except Exception:
        pass

# Funzione per creare file Excel se non esiste
def crea_file_se_non_esiste(file_name):
    if not os.path.exists(file_name):
        wb = Workbook()
        ws = wb.active
        ws.title = "Orari"
        ws.append(["Data", "Reparto", "Entrata", "Uscita", "Ore Lavorate"])
        wb.save(file_name)

def registra_entrata():
    global entrata
    entrata = datetime.now()
    salva_stato()
    status_label.config(text=f"Entrata registrata: {entrata.strftime('%H:%M:%S')}")

def registra_uscita():
    global entrata
    if not entrata:
        status_label.config(text="Errore: devi prima registrare un'entrata!")
        return

    uscita = datetime.now()
    ore = round((uscita - entrata).total_seconds() / 3600, 2)

    if ore < 0:
        status_label.config(text="Errore: l'uscita non può essere prima dell'entrata!")
        return

    try:
        file_name = get_file_name()
        crea_file_se_non_esiste(file_name)
        wb = load_workbook(file_name)
        ws = wb.active
        ws.append([
            datetime.now().strftime("%Y-%m-%d"),
            reparto_var.get(),
            entrata.strftime("%H:%M"),
            uscita.strftime("%H:%M"),
            ore
        ])
        wb.save(file_name)

        status_label.config(text=f"Uscita registrata: {uscita.strftime('%H:%M')} - {ore}h")
        entrata = None
        salva_stato()
    except Exception as e:
        status_label.config(text=f"Errore nel salvataggio: {e}")

def inserisci_manual():
    data = data_entry.get()
    entrata_m = entrata_entry.get()
    uscita_m = uscita_entry.get()
    reparto_m = reparto_manual.get()

    if not data or not entrata_m or not uscita_m or not reparto_m:
        status_label.config(text="Errore: compila tutti i campi!")
        return

    try:
        entrata_dt = datetime.strptime(f"{data} {entrata_m}", "%Y-%m-%d %H:%M")
        uscita_dt = datetime.strptime(f"{data} {uscita_m}", "%Y-%m-%d %H:%M")
        ore = round((uscita_dt - entrata_dt).total_seconds() / 3600, 2)

        if ore < 0:
            status_label.config(text="Errore: l'uscita non può essere prima dell'entrata!")
            return

        file_name = get_file_name(data)
        crea_file_se_non_esiste(file_name)
        wb = load_workbook(file_name)
        ws = wb.active
        ws.append([data, reparto_m, entrata_m, uscita_m, ore])
        wb.save(file_name)

        status_label.config(text=f"Turno manuale inserito: {data} {reparto_m} - {ore}h")
    except ValueError:
        status_label.config(text="Errore: formato data/ora non valido (usa YYYY-MM-DD e HH:MM)")
    except Exception as e:
        status_label.config(text=f"Errore: {e}")

def genera_grafico_mensile():
    import glob

    # Trova tutti i file di orari nella directory corrente
    files = glob.glob("orari_lavoro_*.xlsx")

    if not files:
        status_label.config(text="Nessun file di orari trovato!")
        return

    # Calcolo ore per mese aggregando tutti i file
    mesi = {}
    for file in files:
        try:
            wb = load_workbook(file)
            ws = wb.active

            for row in ws.iter_rows(min_row=2, values_only=True):
                if not row[0] or not row[4]:
                    continue

                data, ore = row[0], row[4]
                # Gestisce sia stringhe che oggetti datetime da Excel
                if isinstance(data, str):
                    mese = datetime.strptime(data, "%Y-%m-%d").strftime("%B %Y")
                else:
                    mese = data.strftime("%B %Y")
                mesi[mese] = mesi.get(mese, 0) + ore
            wb.close()
        except Exception:
            continue

    if not mesi:
        status_label.config(text="Nessun dato da elaborare!")
        return

    # Crea un file di riepilogo mensile
    summary_file = "orari_riepilogo_mensile.xlsx"
    wb_summary = Workbook()
    ws_summary = wb_summary.active
    ws_summary.title = "Riepilogo Mensile"
    ws_summary.append(["Mese", "Ore Totali"])

    for mese, ore in sorted(mesi.items()):
        ws_summary.append([mese, ore])

    # Grafico comparativo
    chart = BarChart()
    chart.title = "Ore lavorate per mese"
    chart.x_axis.title = "Mese"
    chart.y_axis.title = "Ore Totali"

    data = Reference(ws_summary, min_col=2, min_row=2, max_row=ws_summary.max_row)
    cats = Reference(ws_summary, min_col=1, min_row=2, max_row=ws_summary.max_row)
    chart.add_data(data, titles_from_data=False)
    chart.set_categories(cats)

    ws_summary.add_chart(chart, "D2")
    wb_summary.save(summary_file)
    status_label.config(text=f"Grafico mensile salvato in {summary_file}!")

# Carica stato all'avvio
carica_stato()

# GUI
root = tk.Tk()
root.title("Tracciamento Orari Avanzato")

# Selezione reparto automatico
ttk.Label(root, text="Seleziona reparto:").pack(pady=5)
reparto_var = tk.StringVar(value="Magazzino")
ttk.Combobox(root, textvariable=reparto_var, values=["Magazzino", "Produzione", "Ufficio"]).pack(pady=5)

ttk.Button(root, text="Entrata", command=registra_entrata).pack(pady=5)
ttk.Button(root, text="Uscita", command=registra_uscita).pack(pady=5)

# Inserimento manuale
ttk.Label(root, text="--- Inserimento manuale ---").pack(pady=10)

ttk.Label(root, text="Data (YYYY-MM-DD):").pack()
data_entry = ttk.Entry(root)
data_entry.insert(0, "2025-11-21")
data_entry.pack(pady=5)

ttk.Label(root, text="Orario entrata (HH:MM):").pack()
entrata_entry = ttk.Entry(root)
entrata_entry.insert(0, "08:00")
entrata_entry.pack(pady=5)

ttk.Label(root, text="Orario uscita (HH:MM):").pack()
uscita_entry = ttk.Entry(root)
uscita_entry.insert(0, "16:00")
uscita_entry.pack(pady=5)

ttk.Label(root, text="Reparto:").pack()
reparto_manual = ttk.Combobox(root, values=["Magazzino", "Produzione", "Ufficio"])
reparto_manual.set("Magazzino")
reparto_manual.pack(pady=5)

ttk.Button(root, text="Inserisci Manualmente", command=inserisci_manual).pack(pady=5)
ttk.Button(root, text="Genera Grafico Mensile", command=genera_grafico_mensile).pack(pady=5)

status_label = ttk.Label(root, text="Pronto")
status_label.pack(pady=10)

# Mostra stato entrata se caricato
if entrata:
    status_label.config(text=f"Entrata in corso da: {entrata.strftime('%H:%M:%S del %d/%m/%Y')}")

root.mainloop()